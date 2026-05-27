#AUTORES: Gabriel Aparicio y Jeaninne Aparicio – Proyecto IoT Puingenierías
#FECHA DE INICIO: 03-09-2024
#LICENCIA: Trabajo Académico – Distribución Restringida
#ÚLTIMA ACTUALIZACIÓN: 15-12-2025

from django.core.mail import send_mail
import os
from django.views.decorators.csrf import csrf_exempt
import json
from django.http import JsonResponse
from django import forms
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponse
from django.conf import settings
from django.views.generic import TemplateView, ListView, DetailView, CreateView, UpdateView, DeleteView
from openpyxl import Workbook
from openpyxl.styles import Font
import datetime
from reportlab.lib.pagesizes import A4
from decimal import Decimal
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from django.db.models import Sum
from io import BytesIO
from django.core.mail import EmailMessage
from mediciones.models import Medicion
from django.template.loader import render_to_string
from django.contrib.auth import get_user_model
from .forms import CotizacionForm, OrdenTrabajo ,HerramientaForm, RegistroUsoHerramientaForm, OrdenTrabajoForm, PresupuestoForm, OrdenCompraForm, FacturaForm, CondicionTallerForm, ReporteGeneradoForm, TrabajadorForm, DisponibilidadForm
from .models import Cotizacion,ComentarioOT, Alerta ,Herramienta,HistorialRFID, RegistroUsoHerramienta, OrdenTrabajo, Presupuesto, OrdenCompra, Factura, CondicionTaller, ReporteGenerado, ComentarioOT, Trabajador
from .utils import enviar_notificacion_cotizacion
from django.urls import reverse_lazy
from datetime import timedelta
import jwt
from Sistema_proyecto.utils import staff_required,admin_required_redirect
from django.utils.decorators import method_decorator
from django.urls import reverse
from .utils import enviar_presupuesto_email
from io import BytesIO
from django.core.files.base import ContentFile
from django.views import View
from django.core.files import File
from Sistema_proyecto.models import Herramienta
from Sistema_proyecto.utils import calcular_costo_energia
from django.db.models import Sum
from django.db import models
from sensor.models import Sensor
from django.db.models import Avg, Sum
from decimal import Decimal
from django.db.models import Exists, OuterRef
from django.shortcuts import render
from django.utils import timezone
from .prediccion import predecir_consumo_y_costo
from django.db.models import Avg, Prefetch
import joblib
import numpy as np

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MODELO_PATH = os.path.join(BASE_DIR, 'ml_models', 'modelo_energia.pkl')
SCALER_PATH = os.path.join(BASE_DIR, 'ml_models', 'scaler.pkl')
TARIFA_KWH = 150  # Pesos chilenos por kWh (ajusta a tu tarifa real)
CONFIG_PATH = os.path.join(settings.BASE_DIR, 'ml_models', 'config.pkl')

User = get_user_model()

def inicio(request):
    """Redirige siempre al home público."""
    return redirect('home')

class HomeView(TemplateView):
    template_name = "Sistema_proyecto/publico/home.html"

class QuienesSomosView(TemplateView):
    template_name = "Sistema_proyecto/publico/quienes_somos.html"

class ServiciosView(TemplateView):
    template_name = "Sistema_proyecto/publico/servicios.html"

class UbicacionView(TemplateView):
    template_name = "Sistema_proyecto/publico/ubicacion.html"

class ContactoView(TemplateView):
    template_name= "Sistema_proyecto/publico/contacto.html"

def enviar_correo(request):
    if request.method == "POST":
        nombre = request.POST.get("nombre")
        empresa = request.POST.get("empresa")
        servicio = request.POST.get("servicio")
        correo = request.POST.get("correo")
        telefono = request.POST.get("telefono")
        descripcion = request.POST.get("descripcion")
        archivo = request.FILES.get("archivo")

        cot = Cotizacion.objects.create(
            usuario=None,               
            nombre=nombre,
            empresa=empresa,
            correo=correo,
            telefono=telefono,
            servicio= servicio,
            descripcion=descripcion,
            archivo=archivo,
            fecha_entrega=timezone.now().date(),   
            estado="pendiente",
        )

        template = render_to_string(
            'Sistema_proyecto/publico/correo.html',
            {
                'nombre': nombre,
                'empresa':empresa,
                'servicio':servicio,
                'correo': correo,
                'telefono': telefono,
                'mensaje': descripcion
                
            }
        )

        email = EmailMessage(
            subject=f"Solicitud de cotización de {nombre}",
            body=template,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=['yany.china@gmail.com'],
        )
        email.content_subtype = "html"

        if archivo:
            email.attach(archivo.name, archivo.read(), archivo.content_type)

        email.send()

        messages.success(request, "Tu cotización fue enviada y registrada correctamente.")
        return redirect("cotizacion_crear")

    return redirect("cotizacion_crear")

@admin_required_redirect
def reenviar_presupuesto(request, pk):
    presupuesto = get_object_or_404(Presupuesto, id=pk)

    enviar_presupuesto_email(presupuesto)  
    messages.success(request, "El presupuesto fue reenviado correctamente.")
    return redirect("lista_presupuestos")



def login_view(request):
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            return redirect("home_admin")
        else:
            messages.error(request, "Usuario o contraseña incorrectos.")
            return redirect("login")

    return render(request, "Sistema_proyecto/login.html")


def logout_view(request):
    logout(request)
    messages.info(request, "Sesión cerrada correctamente.")
    return redirect("login")

def registro_view(request):

    if request.method == "POST":

        rut = request.POST.get("rut")
        dv = request.POST.get("dv")
        username = request.POST.get("username")
        telefono = request.POST.get("telefono")
        password = request.POST.get("password")
        password2 = request.POST.get("password2")

        # Validar password
        if password != password2:
            messages.error(request, "Las contraseñas no coinciden.")
            return redirect("registro")

        # Validar RUT
        if not rut or not dv:
            messages.error(request, "Debe ingresar el RUT completo.")
            return redirect("registro")

        rut_final = f"{rut}-{dv}"

        # Validar usuario duplicado
        if User.objects.filter(username=username).exists():
            messages.error(request, "El nombre de usuario ya existe.")
            return redirect("registro")

        # Validar RUT duplicado
        if User.objects.filter(rut=rut_final).exists():
            messages.error(request, "El RUT ya está registrado.")
            return redirect("registro")

        # Crear usuario administrador
        user = User(
            username=username,
            telefono=telefono,
            rut=rut_final,
            is_staff=True,       
            is_superuser=False   
        )
        user.set_password(password)
        user.save()

        messages.success(request, "Cuenta de administradora creada correctamente.")
        return redirect("login")

    return render(request, "Sistema_proyecto/registro.html")


@admin_required_redirect
def dashboard_admin(request):
    limite = timezone.now() - timedelta(minutes=1)

    total_sensores = (
        Sensor.objects.filter(mediciones__fecha__gte=limite)
        .distinct()
        .count()
    )

    contexto = {
        "total_mediciones": Medicion.objects.count(),
        "total_sensores": total_sensores,
        "total_cotizaciones": Cotizacion.objects.count(),
        "total_ot": OrdenTrabajo.objects.count(),
    }

    return render(request, "Sistema_proyecto/administrador/home_admin.html", contexto)

@login_required
def informe_energia(request):
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")

    datos = Medicion.objects.all().order_by("-fecha")

    if fecha_inicio and fecha_fin:
        try:
            inicio = datetime.datetime.strptime(fecha_inicio, "%Y-%m-%d")
            fin = datetime.datetime.strptime(fecha_fin, "%Y-%m-%d")
            fin = fin.replace(hour=23, minute=59, second=59)
            datos = datos.filter(fecha__range=(inicio, fin))
        except:
            messages.error(request, "Formato de fecha inválido.")

    ordenes = OrdenTrabajo.objects.all().order_by('-id')

    contexto = {
        "datos": datos,
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "ordenes": ordenes,
    }

    return render(request, "Sistema_proyecto/administrador/informe.html", contexto)


@login_required
def exportar_informe_excel(request):
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")

    wb = Workbook()
    ws = wb.active
    ws.title = "Informe Energía"

    ws["A1"] = "Fecha"
    ws["B1"] = "Corriente (A)"
    ws["C1"] = "Energía (kWh)"
    ws["D1"] = "Costo (CLP)"
    ws["E1"] = "Temperatura"
    ws["F1"] = "Humedad"
    ws["G1"] = "Vibración"
    ws["H1"] = "Orden de Trabajo"


    for c in ws["1:1"]:
        c.font = Font(bold=True)

    qs = Medicion.objects.order_by("fecha")

    if fecha_inicio and fecha_fin:
        qs = qs.filter(fecha__range=[fecha_inicio, fecha_fin])

    fila = 2
    for m in qs:
        ws[f"A{fila}"] = m.fecha.strftime("%Y-%m-%d %H:%M")
        ws[f"B{fila}"] = m.corriente
        ws[f"C{fila}"] = m.energia
        ws[f"D{fila}"] = m.costo
        ws[f"E{fila}"] = m.temperatura
        ws[f"F{fila}"] = m.humedad
        ws[f"G{fila}"] = m.vibracion

        fila += 1

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nombre = f"informe_energia_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"

    reporte = ReporteGenerado.objects.create(
        tipo="Informe Energético",
        
    )
    reporte.archivo.save(nombre, ContentFile(buffer.getvalue()))

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename=\"{nombre}\"'
    return response


@admin_required_redirect
def exportar_ot_excel(request, pk):
    ot = get_object_or_404(OrdenTrabajo, id=pk)

    wb = Workbook()
    ws = wb.active
    ws.title = f"OT-{ot.id}"

    ws.append(["ID OT", "Cliente", "Servicio", "Fecha", "Energía (kWh)", "Corriente (A)", "Costo Energía"])
    for c in ws["1:1"]:
        c.font = Font(bold=True)

    # Corriente promedio de mediciones asociadas
    from django.db.models import Avg
    corriente = ot.medicion_set.aggregate(Avg('corriente'))['corriente__avg'] or 0

    ws.append([
        ot.id,
        ot.cliente,
        ot.servicio,
        ot.fecha_creacion.strftime("%d/%m/%Y %H:%M"),
        ot.energia_total_kwh or 0,
        round(corriente, 2),
        float(ot.costo_energia or 0),
    ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nombre = f"OT_{ot.id}_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre}"'
    return response


@login_required
def exportar_mediciones_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.append(["Fecha", "Corriente", "Energía", "Costo (CLP)", "Temperatura", "Humedad", "Vibración"])

    for m in Medicion.objects.order_by("fecha"):
        ws.append([
            m.fecha.strftime("%Y-%m-%d %H:%M"),
            m.corriente,
            m.energia,
            m.costo,
            m.temperatura,
            m.humedad,
            m.vibracion
        ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nombre = f"mediciones_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"

    reporte = ReporteGenerado.objects.create(
        tipo="Reporte de Mediciones",
    )
    reporte.archivo.save(nombre, ContentFile(buffer.getvalue()))

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre}"'
    return response


@login_required
def limpiar_mediciones(request):
    Medicion.objects.all().delete()
    messages.success(request, "Todas las mediciones fueron eliminadas.")
    return redirect("home_admin")


@login_required
def respaldar_y_limpiar(request):
    fecha_limite = timezone.now() - datetime.timedelta(days=7)
    qs = Medicion.objects.filter(fecha__lt=fecha_limite)

    if not qs.exists():
        messages.info(request, "No hay mediciones antiguas para respaldar.")
        return redirect("home_admin")

    wb = Workbook()
    ws = wb.active
    ws.append(["Fecha", "Corriente", "Energía", "Costo (CLP)", "Temperatura", "Humedad", "Vibración"])

    for m in qs:
        ws.append([m.fecha.strftime("%Y-%m-%d"), m.corriente, m.energia, m.costo, m.temperatura, m.humedad, m.vibracion])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nombre = f"respaldo_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"

    reporte = ReporteGenerado.objects.create(
        tipo="Respaldo Automático",
    )
    reporte.archivo.save(nombre, ContentFile(buffer.getvalue()))

    qs.delete()

    messages.success(request, "Respaldo generado y datos eliminados.")

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename=\"{nombre}\"'
    return response

DASHBOARD_ID = 1
@admin_required_redirect
def metabase_embed(request):
    
    params = {
        # "cliente_id": 123,
        # "fecha_inicio": "2025-10-01",
        # "fecha_fin": "2025-10-31",
    }

    payload = {
        "resource": {"dashboard": DASHBOARD_ID},
        "params": params,
        "exp": datetime.datetime.utcnow() + datetime.timedelta(minutes=10)  
    }

    token = jwt.encode(payload, settings.METABASE_EMBEDDING_SECRET, algorithm="HS256")
    

    iframe_url = f"{settings.METABASE_SITE_URL}/embed/dashboard/{token}#bordered=true&titled=true"

    return render(request, "Sistema_proyecto/reportes/metabase_embed.html", {"iframe_url": iframe_url})

def cotizacion_crear_view(request):
    if request.method == "POST":
        form = CotizacionForm(request.POST, request.FILES)

        if form.is_valid():
            cot = form.save(commit=False)
            cot.usuario = None
            cot.fecha_entrega = timezone.now().date()
            cot.estado = "pendiente"
            cot.save()

            # URLs absolutas para aceptar/rechazar
            aceptar_url = request.build_absolute_uri(
                reverse('aceptar_cotizacion', args=[cot.id])
            )
            rechazar_url = request.build_absolute_uri(
                reverse('rechazar_cotizacion', args=[cot.id])
            )

            template = render_to_string(
                'Sistema_proyecto/publico/correo.html',
                {
                    'nombre':       cot.nombre,
                    'empresa':      cot.empresa,
                    'servicio':     cot.servicio,
                    'correo':       cot.correo,
                    'telefono':     cot.telefono,
                    'mensaje':      cot.descripcion,
                    'aceptar_url':  aceptar_url,
                    'rechazar_url': rechazar_url,
                }
            )

            email = EmailMessage(
                subject=f"Solicitud de cotización de {cot.nombre}",
                body=template,
                from_email=settings.DEFAULT_FROM_EMAIL,
                to=['yany.china@gmail.com'],
            )
            email.content_subtype = "html"

            if cot.archivo:
                cot.archivo.seek(0)
                email.attach(
                    cot.archivo.name,
                    cot.archivo.read(),
                    'application/octet-stream'
                )

            try:
                email.send(fail_silently=False)
                messages.success(request, "Tu cotización fue enviada y registrada correctamente.")
            except Exception as e:
                messages.warning(request, "Cotización registrada, pero hubo un problema al enviar el correo.")
                print(f"❌ Error email: {e}")

            return redirect("cotizacion_crear")

        else:
            print("❌ Errores del form:", form.errors.as_json())
            messages.error(request, "Por favor corrige los errores del formulario.")

    else:
        servicio_inicial = request.GET.get("servicio", "")
        form = CotizacionForm(initial={"servicio": servicio_inicial})

    return render(request, "Sistema_proyecto/publico/cotizacion_form.html", {"form": form})



@admin_required_redirect
def perfil_administrador(request):
    admin = request.user
    return render(request, 'Sistema_proyecto/administrador/perfil_admin.html', {'admin': admin})


@admin_required_redirect
def reportes_administrador(request):
    context = {
        "total_mediciones": Medicion.objects.count(),
        "total_herramientas": Herramienta.objects.count(),
        "total_ot": OrdenTrabajo.objects.count(),
        "reportes": ReporteGenerado.objects.all().order_by("-fecha"),
    }

    return render(request, "Sistema_proyecto/reportes/reportes_admin.html", context)

@admin_required_redirect
def cotizaciones_admin(request):

    cotizaciones = Cotizacion.objects.all().order_by('-id')

    return render(
        request,
        "Sistema_proyecto/administrador/cotizacion_lista.html",
        {"cotizaciones": cotizaciones}
    )

@staff_required
def aceptar_cotizacion(request, id):
    cotizacion = get_object_or_404(Cotizacion, id=id)
    cotizacion.estado = "aprobada"
    cotizacion.save()

    # Enviar CORREO al cliente
    enviar_notificacion_cotizacion(cotizacion, "aprobada")

    return redirect("cotizacion_detalle", pk=id)



@staff_required
def rechazar_cotizacion(request, id):
    cotizacion = get_object_or_404(Cotizacion, id=id)
    cotizacion.estado = "rechazada"
    cotizacion.save()

    messages.info(request, "La cotización fue rechazada.")
    return redirect("cotizacion_detalle", pk=id)


@method_decorator(admin_required_redirect, name="dispatch")
class DetalleCotizacionView(  DetailView):
    model = Cotizacion
    template_name = "Sistema_proyecto/administrador/cotizacion_detalle.html"
    context_object_name = "cotizacion"

    
@admin_required_redirect
def exportar_cotizaciones_pdf(request):
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    p.setFont("Helvetica-Bold", 16)
    p.drawString(2 * cm, height - 2 * cm, "Reporte de Cotizaciones - P&U Ingenierías")

    p.setFont("Helvetica-Bold", 11)
    y = height - 3.5 * cm
    p.drawString(1.5 * cm, y, "ID")
    p.drawString(3.0 * cm, y, "Nombre")
    p.drawString(7.5 * cm, y, "Correo")
    p.drawString(12.5 * cm, y, "Servicio")
    p.drawString(17 * cm, y, "Estado")

    y -= 0.7 * cm
    p.setFont("Helvetica", 10)

    cotizaciones = Cotizacion.objects.all().order_by('-id')

    for cot in cotizaciones:
        if y < 2 * cm:
            p.showPage()
            y = height - 3.5 * cm
        p.drawString(1.5 * cm, y, str(cot.id))
        p.drawString(3.0 * cm, y, cot.nombre[:18])
        p.drawString(7.5 * cm, y, cot.correo[:20])
        p.drawString(12.5 * cm, y, cot.servicio[:15])
        p.drawString(17 * cm, y, cot.get_estado_display())
        y -= 0.6 * cm

    p.save()

    pdf_value = buffer.getvalue()
    nombre = f"reporte_cotizaciones_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"

    reporte = ReporteGenerado.objects.create(
        tipo="Reporte de Cotizaciones",
    )
    reporte.archivo.save(nombre, ContentFile(pdf_value))

    buffer.close()

    response = HttpResponse(pdf_value, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{nombre}"'
    return response


@method_decorator(admin_required_redirect, name="dispatch")
class HerramientaListView(ListView):
    model = Herramienta
    template_name = "Sistema_proyecto/herramientas/lista.html"
    context_object_name = "herramientas"

    def get_queryset(self):
        queryset = super().get_queryset()

        uso_activo = RegistroUsoHerramienta.objects.filter(
            herramienta=OuterRef("pk"),
            fecha_fin__isnull=True
        )

        return queryset.annotate(
            esta_en_uso=Exists(uso_activo)
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        herramientas = context["herramientas"]
        ahora = timezone.now()

        
        # Pre-cargar datos relacionados
        
        registros_por_herramienta = {
            h.id: list(
                RegistroUsoHerramienta.objects.filter(herramienta=h)
            )
            for h in herramientas
        }

        rfid_por_herramienta = {
            h.id: {
                "entrada": HistorialRFID.objects.filter(
                    herramienta=h, accion="entrada"
                ).order_by("-fecha").first(),
                "salida": HistorialRFID.objects.filter(
                    herramienta=h, accion="salida"
                ).order_by("-fecha").first(),
            }
            for h in herramientas
        }

        
        #  HORAS ACUMULADAS
        
        for h in herramientas:
            total_segundos = 0

            for r in registros_por_herramienta.get(h.id, []):
                if not r.fecha_inicio:
                    continue

                fin = r.fecha_fin if r.fecha_fin else ahora
                total_segundos += (fin - r.fecha_inicio).total_seconds()

            h.horas_acumuladas = round(total_segundos / 3600, 2)

        
        #  ÚLTIMA SACADA / ENTREGA
        
        for h in herramientas:
            entrada = rfid_por_herramienta[h.id]["entrada"]
            salida = rfid_por_herramienta[h.id]["salida"]

            h.ultima_sacada = entrada.fecha if entrada else None
            h.ultima_entrega = salida.fecha if salida else None

        return context
    
    
@method_decorator(admin_required_redirect, name="dispatch")
class HerramientaCreateView(  CreateView):
    model = Herramienta
    form_class = HerramientaForm
    template_name = "Sistema_proyecto/herramientas/crear.html"
    success_url = reverse_lazy("lista_herramientas")

@method_decorator(admin_required_redirect, name="dispatch")
class HerramientaDetailView(DetailView):
    model = Herramienta
    template_name = "Sistema_proyecto/herramientas/detalle.html"
    context_object_name = "herramienta"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        herramienta = self.get_object()

        # Obtener historial RFID asociado
        historial = HistorialRFID.objects.filter(
            herramienta=herramienta
        ).order_by("-fecha")

        context["historial"] = historial
        return context


@method_decorator(admin_required_redirect, name="dispatch")
class HerramientaUpdateView(  UpdateView):
    model = Herramienta
    form_class = HerramientaForm
    template_name = "Sistema_proyecto/herramientas/editar.html"
    success_url = reverse_lazy("lista_herramientas")

@method_decorator(admin_required_redirect, name="dispatch")
class HerramientaDeleteView(  DeleteView):
    model = Herramienta
    template_name = "Sistema_proyecto/herramientas/eliminar.html"
    success_url = reverse_lazy("lista_herramientas")


@csrf_exempt
def registrar_rfid(request):
    if request.method != "POST":
        return JsonResponse({"error": "Método no permitido"}, status=405)

    data = json.loads(request.body.decode("utf-8"))

    uid = data.get("uid")
    sensor_id = data.get("sensor_id")

    if not uid or not sensor_id:
        return JsonResponse({"error": "Datos incompletos"}, status=400)

    sensor = get_object_or_404(Sensor, id=sensor_id)
    herramienta = get_object_or_404(Herramienta, uid=uid)

    
    # DETERMINAR ACCIÓN
    
    if herramienta.estado == "disponible":
        accion = "salida"
        herramienta.estado = "en_uso"
    else:
        accion = "entrada"
        herramienta.estado = "disponible"

    herramienta.save()

    
    # REGISTRAR HISTORIAL RFID
    
    HistorialRFID.objects.create(
        herramienta=herramienta,
        sensor=sensor,
        uid=uid,
        accion=accion
    )

    return JsonResponse({
            "status": "ok",
            "accion": accion,                
            "herramienta": herramienta.nombre,
            "estado": herramienta.estado
        })


@method_decorator(admin_required_redirect, name="dispatch")
class RegistroUsoListView(  ListView):
    model = RegistroUsoHerramienta
    template_name = "Sistema_proyecto/herramientas/uso_lista.html"
    context_object_name = "registros"

    


@method_decorator(admin_required_redirect, name="dispatch")
class RegistroUsoCreateView(CreateView):
    model = RegistroUsoHerramienta
    form_class = RegistroUsoHerramientaForm
    template_name = "Sistema_proyecto/herramientas/uso_crear.html"
    success_url = reverse_lazy("lista_registros_uso")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["hoy"] = timezone.localtime().strftime("%Y-%m-%dT%H:%M")
        return context

    def form_valid(self, form):
        uso = form.save(commit=False)

        #  ASOCIAR A LA OT ACTIVA
        ot = OrdenTrabajo.objects.filter(
            herramientas=uso.herramienta,
            estado="en_proceso"
        ).last()

        if not ot:
            form.add_error(None, "No hay una OT activa para esta herramienta.")
            return self.form_invalid(form)

        uso.orden_trabajo = ot
        uso.energia_consumida = 0
        uso.save()

        messages.success(self.request, "Uso registrado correctamente.")
        return super().form_valid(form)


@method_decorator(admin_required_redirect, name="dispatch")
class RegistroUsoDetailView(  DetailView):
    model = RegistroUsoHerramienta
    template_name = "Sistema_proyecto/herramientas/uso_detalle.html"
    context_object_name = "registro"

    


@method_decorator(admin_required_redirect, name="dispatch")
class RegistroUsoUpdateView(  UpdateView):
    model = RegistroUsoHerramienta
    form_class = RegistroUsoHerramientaForm
    template_name = "Sistema_proyecto/herramientas/uso_editar.html"
    success_url = reverse_lazy("lista_registros_uso")
    def form_valid(self, form):
        uso = form.save(commit=False)

        if uso.fecha_inicio and uso.fecha_fin:
            horas = (uso.fecha_fin - uso.fecha_inicio).total_seconds() / 3600

            # Solo si la herramienta tiene sensor SCT-013
            if uso.herramienta.sensor and uso.herramienta.sensor.tipo == "SCT-013":
                CONSUMO_ESTIMADO = 1.0  
                uso.energia_consumida = round(horas * CONSUMO_ESTIMADO, 3)
            else:
                uso.energia_consumida = 0

        uso.save()
        return super().form_valid(form)


    

@method_decorator(admin_required_redirect, name="dispatch")
class RegistroUsoDeleteView(  DeleteView):
    model = RegistroUsoHerramienta
    template_name = "Sistema_proyecto/herramientas/uso_eliminar.html"
    success_url = reverse_lazy("lista_registros_uso")

        

@method_decorator(admin_required_redirect, name="dispatch")
class OrdenTrabajoListView(   ListView):
    model = OrdenTrabajo
    template_name = "Sistema_proyecto/ot/lista.html"
    context_object_name = "ordenes" 
    ordering = ["-fecha_creacion"]
    raise_exception = False
    login_url = "/login/"


 

@admin_required_redirect
def crear_ot_desde_presupuesto(request, pk):
    presupuesto = get_object_or_404(Presupuesto, id=pk)
    oc = OrdenCompra.objects.filter(presupuesto=presupuesto).latest("id")

    if OrdenTrabajo.objects.filter(presupuesto=presupuesto).exists():
        messages.warning(request, "Ya existe una orden de trabajo para este presupuesto.")
        ot = OrdenTrabajo.objects.get(presupuesto=presupuesto)
        return redirect("detalle_ot", pk=ot.id)

    ot = OrdenTrabajo.objects.create(
        cliente=presupuesto.cliente,
        servicio=presupuesto.servicio,
        descripcion=presupuesto.descripcion,
        presupuesto=presupuesto,
        orden_compra=oc,
    )

    ot.numero_ot = f"OT-{presupuesto.id}"
    ot.save()

    
    #  ASOCIAR HERRAMIENTA A LA OT
    
    if presupuesto.herramienta:
        ot.herramientas.add(presupuesto.herramienta)

    messages.success(request, "Orden de trabajo generada correctamente.")
    return redirect("detalle_ot", pk=ot.id)




@method_decorator(admin_required_redirect, name="dispatch")
class OrdenTrabajoDetailView(DetailView):
    model = OrdenTrabajo
    template_name = "Sistema_proyecto/ot/detalle.html"
    context_object_name = "ot"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        ot = self.object

        
        #  REGISTROS DE USO (ENERGÍA REAL)
        
        registros = RegistroUsoHerramienta.objects.filter(
            orden_trabajo=ot,
            herramienta__in=ot.herramientas.all()
        ).order_by("-fecha_inicio")

        energia_raw = registros.aggregate(
            total=Sum("energia_consumida")
        )["total"]

        energia_total = (Decimal(str(energia_raw)).quantize(Decimal("0.001"))if energia_raw else Decimal("0.000"))

        costo_kwh = Decimal("160")
        costo_energia = (energia_total * costo_kwh).quantize(Decimal("0.1"))

        
        #  CONDICIONES AMBIENTALES (IoT)
        
        herramientas = ot.herramientas.all()
        sensores = [h.sensor for h in herramientas if h.sensor]

        mediciones = Medicion.objects.filter(
            sensor__in=sensores
        )
        condiciones = mediciones.aggregate(
            temp_avg=Avg("temperatura"),
            hum_avg=Avg("humedad"),
            vib_avg=Avg("vibracion")
        )

        #  NORMALIZAR None → 0
        temperatura = round(condiciones["temp_avg"], 1) if condiciones["temp_avg"] is not None else 0
        humedad = round(condiciones["hum_avg"], 1) if condiciones["hum_avg"] is not None else 0
        vibracion = round(condiciones["vib_avg"], 2) if condiciones["vib_avg"] is not None else 0

        
        #  DURACIÓN OT
        
        if ot.fecha_inicio and ot.fecha_termino:
            duracion = (ot.fecha_termino - ot.fecha_inicio).total_seconds() / 3600
        else:
            duracion = 0

        duracion = round(duracion, 2)

        
        #  ÍNDICE DE EFICIENCIA
        
        if duracion > 0:
            indice_eficiencia = round(
                (Decimal(duracion) * Decimal("1.2")) /
                (energia_total + Decimal("0.01")) * 100,
                1
            )
        else:
            indice_eficiencia = 0

        
        #  ALERTAS IOT
        
        alertas_iot = Alerta.objects.filter(
            herramienta__in=herramientas
        ).order_by("-fecha")[:10]

        
        #  COLORES SEGÚN ESTADO
        
        mapa_colores = {
            "pendiente": ("#ffcc00", "#777", "#777", "#777", "#777"),
            "asignada": ("#ffcc00", "#0d6efd", "#777", "#777", "#777"),
            "en_proceso": ("#ffcc00", "#0d6efd", "#17a2b8", "#777", "#777"),
            "finalizada": ("#ffcc00", "#0d6efd", "#17a2b8", "#28a745", "#777"),
            "entregada": ("#ffcc00", "#0d6efd", "#17a2b8", "#28a745", "#6610f2"),
        }

        
        #  CONTEXTO FINAL
        
        context.update({
            "registros": registros,

            "energia_total": energia_total,
            "costo_energia": costo_energia,
            "mostrar_consumo": energia_total > 0,

            "temperatura": temperatura,
            "humedad": humedad,
            "vibracion": vibracion,

            "duracion_ot": duracion,
            "indice_eficiencia": indice_eficiencia,

            "alertas_iot": alertas_iot,
            "estado_color": mapa_colores.get(ot.estado),
        })

        return context


@method_decorator(admin_required_redirect, name="dispatch")
class OrdenTrabajoUpdateView(  UpdateView):
    model = OrdenTrabajo
    form_class = OrdenTrabajoForm
    template_name = "Sistema_proyecto/ot/editar.html"
    success_url = reverse_lazy("lista_ot")
    raise_exception = False
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["now"] = timezone.now()
        return context

 

@method_decorator(admin_required_redirect, name="dispatch")
class PresupuestoListView(  ListView):
    model = Presupuesto
    template_name = "Sistema_proyecto/presupuestos/lista.html"
    context_object_name = "presupuestos"
    raise_exception = False

 

@method_decorator(admin_required_redirect, name="dispatch")
class PresupuestoCreateView(CreateView):
    model = Presupuesto
    form_class = PresupuestoForm
    template_name = "Sistema_proyecto/presupuestos/crear.html"
    success_url = reverse_lazy("lista_presupuestos")
    raise_exception = False

    def dispatch(self, request, *args, **kwargs):
        # Capturamos la cotización
        self.cotizacion = get_object_or_404(Cotizacion, id=self.kwargs['cotizacion_id'])
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["cotizacion"] = self.cotizacion

        form = context["form"]  

        herramienta_id = None

        if form.is_bound:
            herramienta_id = form.data.get("herramienta")

      
        # SI NO VIENE DE BOUND, REVISAR INITIAL
      
        if not herramienta_id:
            herramienta_id = form.initial.get("herramienta")

        # Convertir a int si existe
        if herramienta_id:
            try:
                herramienta_id = int(herramienta_id)
            except:
                herramienta_id = None

        # Obtener objeto herramienta
        herramienta_obj = None
        if herramienta_id:
            try:
                herramienta_obj = Herramienta.objects.get(id=herramienta_id)
            except Herramienta.DoesNotExist:
                herramienta_obj = None

      
        # CALCULAR CONSUMO SI HAY HERRAMIENTA
      
        if herramienta_obj:
            context["herramienta_seleccionada"] = True
            context["herramienta_nombre"] = herramienta_obj.nombre

            
            energia_total, costo_energia = calcular_costo_energia(herramienta_obj)

            context["energia_total"] = energia_total
            context["costo_energia"] = costo_energia
        else:
            context["herramienta_seleccionada"] = False
            context["energia_total"] = None
            context["costo_energia"] = None
        monto_base = 0

        if form.is_bound:
            try:
                monto_base = int(form.data.get("monto_base", 0))
            except:
                monto_base = 0
        else:
            monto_base = form.initial.get("monto_estimado", 0)

        context["monto_base"] = monto_base

        # TOTAL FINAL
        try:
            total_final = monto_base + (context.get("costo_energia") or 0)
        except:
            total_final = 0

        context["total_final"] = total_final

        return context



    def get_initial(self):
        initial = super().get_initial()
        initial["cliente"] = self.cotizacion.nombre
        return initial

    def form_valid(self, form):
        presupuesto = form.save(commit=False)
        presupuesto.cotizacion = self.cotizacion

        herramienta = form.cleaned_data.get("herramienta")
        presupuesto.herramienta = herramienta

        monto_base = form.cleaned_data.get("monto_base") or 0

        energia_total = 0
        costo_energia = 0

        if herramienta:
            energia_total, costo_energia = calcular_costo_energia(herramienta)

        presupuesto.energia_total = energia_total
        presupuesto.costo_energia = costo_energia
        presupuesto.monto_estimado = monto_base + costo_energia

        presupuesto.save()

        
        #  Enlaces aceptar / rechazar
        


        aceptar_url = self.request.build_absolute_uri(
            reverse("presupuesto_aceptar", args=[presupuesto.id])
        )

        rechazar_url = self.request.build_absolute_uri(
            reverse("presupuesto_rechazar", args=[presupuesto.id])
        )

        
        #  Enviar correo con template HTML
        


        html_mensaje = render_to_string("Sistema_proyecto/presupuestos/email_presupuesto.html", {
            "presupuesto": presupuesto,
            "cotizacion": self.cotizacion,
            "monto_base": monto_base,
            "costo_energia": costo_energia,
            "total_final": presupuesto.monto_estimado,
            "aceptar_url": aceptar_url,
            "rechazar_url": rechazar_url,
        })

        send_mail(
            "Presupuesto generado para su cotización",
            "",
            settings.DEFAULT_FROM_EMAIL,
            ['yany.china@gmail.com'],
            fail_silently=False,
            html_message=html_mensaje,
        )

        return super().form_valid(form)





 

@method_decorator(admin_required_redirect, name="dispatch")
class PresupuestoDetailView(  DetailView):
    model = Presupuesto
    template_name = "Sistema_proyecto/presupuestos/detalle.html"
    context_object_name = "presupuesto"
    raise_exception = False

 

@method_decorator(admin_required_redirect, name="dispatch")
class PresupuestoUpdateView(  UpdateView):
    model = Presupuesto
    form_class = PresupuestoForm
    template_name = "Sistema_proyecto/presupuestos/editar.html"
    success_url = reverse_lazy("lista_presupuestos")
    raise_exception = False
    def form_valid(self, form):
        presupuesto = form.save(commit=False)

    
        presupuesto.cliente = form.cleaned_data.get("cliente")

        presupuesto.save()
        return super().form_valid(form)



class PresupuestoAceptarView(View):
    def get(self, request, pk):
        presupuesto = get_object_or_404(Presupuesto, pk=pk)

        # Cambiar estado a aceptado
        presupuesto.estado = "aceptado"
        presupuesto.save()

        # Buscar si el presupuesto YA tiene una orden de compra
        oc = OrdenCompra.objects.filter(presupuesto=presupuesto).first()

        # Si NO existe, crear una nueva
        if oc is None:
            oc = OrdenCompra.objects.create(
                presupuesto=presupuesto,
                cotizacion=presupuesto.cotizacion,
                empresa=presupuesto.cotizacion.empresa,
                cliente=presupuesto.cotizacion.nombre,
                monto_total=presupuesto.monto_estimado,
            )

            oc.numero_orden_compra = f"OC-{oc.id:04d}"
            oc.save()

        # Renderizar confirmación
        return render(
            request,
            "Sistema_proyecto/presupuestos/aceptado.html",
            {
                "presupuesto": presupuesto,
                "orden_compra": oc
            }
        )



class PresupuestoRechazarView(View):
    def get(self, request, pk):
        presupuesto = get_object_or_404(Presupuesto, pk=pk)
        presupuesto.estado = "rechazado"
        presupuesto.save()
        return render(request, "Sistema_proyecto/presupuestos/rechazado.html", {
            "presupuesto": presupuesto
        })

@method_decorator(admin_required_redirect, name="dispatch")
class OrdenCompraListView(  ListView):
    model = OrdenCompra
    template_name = "Sistema_proyecto/oc/lista.html"
    context_object_name = "ocs"
    raise_exception = False

 

@method_decorator(admin_required_redirect, name="dispatch")
class OrdenCompraCreateView(CreateView):
    model = OrdenCompra
    form_class = OrdenCompraForm
    template_name = "Sistema_proyecto/oc/crear.html"

    def dispatch(self, request, *args, **kwargs):
        self.presupuesto = get_object_or_404(Presupuesto, id=kwargs["presupuesto_id"])

        #  Si YA existe una OC asociada → redirigir y NO crear otra
        oc_existente = OrdenCompra.objects.filter(presupuesto=self.presupuesto).first()
        if oc_existente:
            messages.warning(request, "Este presupuesto ya tiene una Orden de Compra.")
            return redirect("detalle_oc", pk=oc_existente.id)

        return super().dispatch(request, *args, **kwargs)


    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["presupuesto"] = self.presupuesto
        return ctx

    def form_valid(self, form):
        oc = form.save(commit=False)

        oc.presupuesto = self.presupuesto
        oc.cotizacion = self.presupuesto.cotizacion
        oc.cliente = self.presupuesto.cotizacion.nombre
        oc.empresa = self.presupuesto.cotizacion.empresa
        oc.monto_total = self.presupuesto.monto_estimado

        oc.save()

        oc.numero_orden_compra = f"OC-{oc.id:04d}"
        oc.save()

        self.object = oc

        #  Redirección correcta y obligatoria
        return redirect("detalle_oc", pk=oc.id)





@method_decorator(admin_required_redirect, name="dispatch")
class OrdenCompraDetailView(  DetailView):
    model = OrdenCompra
    template_name = "Sistema_proyecto/oc/detalle.html"
    context_object_name = "oc"

 

@method_decorator(admin_required_redirect, name="dispatch")
class OrdenCompraUpdateView(  UpdateView):
    model = OrdenCompra
    form_class = OrdenCompraForm
    template_name = "Sistema_proyecto/oc/editar.html"
    success_url = reverse_lazy("lista_oc")
    def get_form(self, *args, **kwargs):
        form = super().get_form(*args, **kwargs)

        # Solo ocultar el campo si existe
        if "cotizacion" in form.fields:
            form.fields["cotizacion"].widget = forms.HiddenInput()

        return form

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["oc"] = self.object   
        return context
@admin_required_redirect
def finalizar_ot(request, pk):
    ot = get_object_or_404(OrdenTrabajo, id=pk)

    if ot.estado != "en_proceso":
        messages.error(request, "La OT debe estar en proceso.")
        return redirect("detalle_ot", pk=ot.id)
    
    for herramienta in ot.herramientas.all():
        herramienta.estado = "disponible"
        herramienta.save()

    energia_total = RegistroUsoHerramienta.objects.filter(
        orden_trabajo=ot
    ).aggregate(
        total=Sum("energia_consumida")
    )["total"] or Decimal("0")

    costo_kwh = Decimal("160")
    costo_energia = (Decimal(energia_total) * costo_kwh).quantize(Decimal("0.01"))

    #  CONGELAR EN OT
    ot.energia_total_kwh = float(energia_total)
    ot.costo_energia = costo_energia
    ot.estado = "entregada"
    ot.fecha_termino = timezone.now()
    ot.save()

    messages.success(request, "OT finalizada correctamente.")
    return redirect("detalle_ot", pk=ot.id)




@admin_required_redirect
def entregar_ot(request, pk):
    ot = get_object_or_404(OrdenTrabajo, id=pk)
    ot.estado = "entregada"
    ot.save()

    messages.success(request, "La Orden de Trabajo fue ENTREGADA al cliente.")
    return redirect("detalle_ot", pk)



@admin_required_redirect
def facturar_ot(request, pk):
    ot = get_object_or_404(OrdenTrabajo, id=pk)

    if ot.estado != "entregada":
        messages.error(request, "La OT debe estar ENTREGADA.")
        return redirect("detalle_ot", pk=ot.id)

    # ✅ Si ya existe factura para esta OT, redirigir a la existente
    factura_existente = Factura.objects.filter(orden_trabajo=ot).first()
    if factura_existente:
        return redirect("detalle_factura", pk=factura_existente.id)

    monto_base = ot.presupuesto.monto_estimado
    costo_energia = ot.costo_energia or Decimal("0")

    subtotal = monto_base + costo_energia
    iva = subtotal * Decimal("0.19")
    total = subtotal + iva

    factura = Factura.objects.create(
        numero_factura=f"FAC-{ot.id}",
        orden_trabajo=ot,
        orden_compra=ot.orden_compra,
        monto_neto=monto_base,
        costo_energia=costo_energia,
        iva=iva,
        monto_total=total,
    )

    return redirect("detalle_factura", pk=factura.id)


@admin_required_redirect
def iniciar_ot(request, pk):
    ot = get_object_or_404(OrdenTrabajo, id=pk)

    herramienta = ot.herramienta

    for herramienta in ot.herramientas.all():
        if herramienta.estado != "disponible":
            messages.error(request, "Una herramienta no está disponible.")
            return redirect("detalle_ot", pk=ot.id)

        herramienta.estado = "en_uso"
        herramienta.save()


    ot.estado = "en_proceso"
    ot.fecha_inicio = timezone.now()
    ot.save()

    messages.success(request, "OT iniciada correctamente.")
    return redirect("detalle_ot", pk=ot.id)

class FacturaListView(ListView):
    model = Factura
    template_name = "Sistema_proyecto/facturas/mostrar.html"
    context_object_name = "facturas"

def mostrar_factura(request, pk):
    factura = get_object_or_404(Factura, pk=pk)
    return render(
        request,
        "Sistema_proyecto/facturas/mostrar.html",
        {"factura": factura}
    )


@method_decorator(admin_required_redirect, name="dispatch")
class FacturaDetailView(  DetailView):
    model = Factura
    template_name = "Sistema_proyecto/facturas/detalle.html"
    context_object_name = "factura"



@method_decorator(admin_required_redirect, name="dispatch")
class FacturaListView(  ListView):
    model = Factura
    template_name = "Sistema_proyecto/facturas/lista.html"
    context_object_name = "facturas"
    
@admin_required_redirect
def factura_pdf(request, pk):
    factura = get_object_or_404(Factura, pk=pk)

    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    
    # ENCABEZADO EMPRESA
    
    p.setFont("Helvetica-Bold", 16)
    p.drawString(2*cm, height - 2*cm, "P&U INGENIERÍAS E.I.R.L.")

    p.setFont("Helvetica", 10)
    p.drawString(2*cm, height - 2.7*cm, "Servicios Industriales e IoT")
    p.drawString(2*cm, height - 3.2*cm, "Correo: contacto@puingenierias.cl")
    p.drawString(2*cm, height - 3.7*cm, "Teléfono: +56 9 1234 5678")

    
    # DATOS FACTURA
    
    p.setFont("Helvetica-Bold", 14)
    p.drawRightString(width - 2*cm, height - 2*cm, "FACTURA")

    p.setFont("Helvetica", 10)
    p.drawRightString(width - 2*cm, height - 2.8*cm, f"N°: {factura.numero_factura}")
    p.drawRightString(
        width - 2*cm,
        height - 3.4*cm,
        f"Fecha: {factura.fecha.strftime('%d-%m-%Y %H:%M')}"
    )

     
    # DATOS CLIENTE
    
    y = height - 5*cm
    p.setFont("Helvetica-Bold", 11)
    p.drawString(2*cm, y, "Datos del Cliente")

    p.setFont("Helvetica", 10)
    y -= 0.7*cm
    p.drawString(2*cm, y, f"Cliente: {factura.orden_trabajo.cliente}")
    y -= 0.6*cm
    p.drawString(2*cm, y, f"Orden de Trabajo: OT-{factura.orden_trabajo.id}")

    if factura.orden_compra:
        y -= 0.6*cm
        p.drawString(
            2*cm,
            y,
            f"Orden de Compra: {factura.orden_compra.numero_orden_compra}"
        )

    
    # DETALLE DEL SERVICIO
    
    y -= 1.5*cm
    p.setFont("Helvetica-Bold", 11)
    p.drawString(2*cm, y, "Detalle del Servicio")

    y -= 0.7*cm
    p.setFont("Helvetica-Bold", 10)
    p.drawString(2*cm, y, "Descripción")
    p.drawRightString(width - 2*cm, y, "Monto Neto")

    p.line(2*cm, y - 0.2*cm, width - 2*cm, y - 0.2*cm)

    y -= 0.7*cm
    p.setFont("Helvetica", 10)
    p.drawString(
        2*cm,
        y,
        factura.orden_trabajo.descripcion or "Servicio industrial"
    )
    p.drawRightString(
        width - 2*cm,
        y,
        f"${int(factura.monto_neto):,}".replace(",", ".")
    )

    
    # RESUMEN ENERGÉTICO (IoT)
 
    y -= 1.5*cm
    p.setFont("Helvetica-Bold", 11)
    p.drawString(2*cm, y, "Resumen Energético (IoT)")

    p.setFont("Helvetica", 10)
    y -= 0.7*cm
    p.drawString(
        2*cm,
        y,
        f"Energía total registrada: {factura.costo_energia / 160:.3f} kWh"
    )

    y -= 0.6*cm
    p.drawString(
        2*cm,
        y,
        f"Costo energético (neto): ${int(factura.costo_energia):,}".replace(",", ".")
    )

    
    # RESUMEN DE TOTALES
    
    y -= 1.5*cm
    p.setFont("Helvetica-Bold", 11)
    p.drawRightString(width - 2*cm, y, f"Subtotal Neto: ${int(factura.monto_neto):,}".replace(",", "."))

    y -= 0.6*cm
    p.drawRightString(width - 2*cm, y, f"IVA 19%: ${int(factura.iva):,}".replace(",", "."))

    y -= 0.8*cm
    p.setFont("Helvetica-Bold", 13)
    p.drawRightString(
        width - 2*cm,
        y,
        f"TOTAL A PAGAR: ${int(factura.monto_total):,}".replace(",", ".")
    )

     
    # PIE DE PÁGINA
    #
    p.setFont("Helvetica-Oblique", 8)
    p.drawCentredString(
        width / 2,
        1.5*cm,
        "Documento generado automáticamente por el sistema IoT Puingenierías"
    )

    p.showPage()
    p.save()
    buffer.seek(0)

    return HttpResponse(
        buffer,
        content_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="factura_{factura.numero_factura}.pdf"'
        }
    )

  
@method_decorator(admin_required_redirect, name="dispatch")
class CondicionesTallerView(LoginRequiredMixin,ListView):
    model = Medicion
    template_name = "Sistema_proyecto/taller/condiciones.html"
    context_object_name = "registros"
    ordering = ["-fecha"]

    def get_queryset(self):
        # una medición cada 10 minutos aprox
        return Medicion.objects.all().order_by("-fecha")[:50]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        for r in context["registros"]:
           
            # TEMPERATURA
            if r.temperatura <= 30:
                r.temp_estado = "normal"
            elif r.temperatura <= 40:
                r.temp_estado = "leve"
            else:
                r.temp_estado = "riesgo"

           
            # HUMEDAD
            if r.humedad <= 60:
                r.hum_estado = "normal"
            elif r.humedad <= 75:
                r.hum_estado = "leve"
            else:
                r.hum_estado = "riesgo"

           
            # VIBRACIÓN
            if r.vibracion <= 40:
                r.vib_estado = "normal"
            elif r.vibracion <= 70:
                r.vib_estado = "leve"
            else:
                r.vib_estado = "riesgo"

           
            # ESTADO GENERAL
            if "riesgo" in [r.temp_estado, r.hum_estado, r.vib_estado]:
                r.estado_general = "riesgo"
            elif "leve" in [r.temp_estado, r.hum_estado, r.vib_estado]:
                r.estado_general = "leve"
            else:
                r.estado_general = "normal"

        return context



@method_decorator(admin_required_redirect, name="dispatch")
class ReporteListView(  ListView):
    model = ReporteGenerado
    template_name = "Sistema_proyecto/reportes/lista.html"
    context_object_name = "reportes"



    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        context["total_mediciones"] = Medicion.objects.count()
        context["total_herramientas"] = Herramienta.objects.count()
        context["total_ot"] = OrdenTrabajo.objects.count()
        context["reportes"] = ReporteGenerado.objects.all()

        return context
 

@method_decorator(admin_required_redirect, name="dispatch")
class ComentarioOTCreateView(CreateView):
    model = ComentarioOT
    fields = ["texto"]
    template_name = "Sistema_proyecto/ot/detalle.html"  

    def form_valid(self, form):
        ot_id = self.kwargs["ot_id"]
        orden = OrdenTrabajo.objects.get(id=ot_id)


        comentario = form.save(commit=False)
        comentario.orden = orden
        comentario.autor = self.request.user
        comentario.save()

        return redirect("detalle_ot", pk=ot_id)


@staff_required
def alertas_lista(request):
    ultima = Medicion.objects.order_by('-fecha').first()

    # ← AGREGAR ESTE CHECK
    if ultima is None:
        return render(request, "Sistema_proyecto/administrador/alertas_lista.html", {
            "temp": 0,
            "temp_color": "primary",
            "hum": 0,
            "hum_color": "primary",
            "vib": 0,
            "vib_color": "primary",
            "corr": 0,
            "corr_color": "primary",
            "alertas_detectadas": ["No hay mediciones registradas aún."],
        })

    def get_color(valor, limites):
        bajo, medio = limites
        if valor >= medio:
            return "danger"
        elif valor >= bajo:
            return "warning"
        return "primary"

    alertas_detectadas = []

    if ultima.temperatura > 40:
        alertas_detectadas.append(f"Temperatura alta: {ultima.temperatura}°C")

    if ultima.vibracion > 50:
        alertas_detectadas.append(f"Vibración peligrosa detectada: {ultima.vibracion}")

    if ultima.corriente > 2:
        alertas_detectadas.append(f"Sobrecorriente detectada: {ultima.corriente}A")

    hora_actual = timezone.localtime().hour
    if 22 <= hora_actual or hora_actual < 7:
        if ultima.corriente > 0.2:
            alertas_detectadas.append("Uso de herramienta fuera de horario laboral")

    contexto = {
        "temp": ultima.temperatura,
        "temp_color": get_color(ultima.temperatura, (30, 40)),
        "hum": ultima.humedad,
        "hum_color": get_color(ultima.humedad, (50, 70)),
        "vib": ultima.vibracion,
        "vib_color": get_color(ultima.vibracion, (40, 70)),
        "corr": ultima.corriente,
        "corr_color": get_color(ultima.corriente, (1, 2)),
        "alertas_detectadas": alertas_detectadas,
    }
    return render(request, "Sistema_proyecto/administrador/alertas_lista.html", contexto)


def vista_prediccion(request):
    resultado = None
    error = None

    if request.method == 'POST':
        try:
            voltaje    = float(request.POST.get('voltaje', 220))
            corriente  = float(request.POST.get('corriente', 5))
            potencia   = float(request.POST.get('potencia', 1100))
            ahora      = timezone.now()
            hora       = ahora.hour
            dia_semana = ahora.weekday()
            mes        = ahora.month

            resultado = predecir_consumo_y_costo(
                voltaje, corriente, potencia, hora, dia_semana, mes
            )
        except Exception as e:
            error = f'Error al predecir: {str(e)}'

    return render(request, 'prediccion.html', {
        'resultado': resultado,
        'error': error,
    })


def predecir_consumo_y_costo(potencia, hora, dia_semana, mes, dias=1):
    modelo = joblib.load(MODELO_PATH)
    scaler = joblib.load(SCALER_PATH)
    config = joblib.load(CONFIG_PATH)
    tarifa = config.get('tarifa', 150)

    kwh_estimado = potencia / 1000

    features = np.array([[
        potencia,
        hora,
        dia_semana,
        mes,
        kwh_estimado,
        kwh_estimado,
        kwh_estimado,
        kwh_estimado,
    ]])

    features_scaled = scaler.transform(features)
    kwh_por_hora    = modelo.predict(features_scaled)[0]
    kwh_total       = kwh_por_hora * 24 * dias
    costo_total     = kwh_total * tarifa

    return {
        'kwh':  round(float(kwh_total), 4),
        'costo': round(float(costo_total), 2),
    }


def api_prediccion(request):
    try:
        dias       = int(request.GET.get('dias', 30))
        ahora      = timezone.now()
        hora       = ahora.hour
        dia_semana = ahora.weekday()
        mes        = ahora.month

        from mediciones.models import Medicion
        ultima = Medicion.objects.order_by('-fecha').first()
        potencia = float(ultima.potencia) if ultima else 1100.0

        # Generar predicción día por día
        predicciones = []
        for i in range(dias):
            fecha_pred = ahora + timezone.timedelta(days=i+1)
            resultado = predecir_consumo_y_costo(
                potencia=potencia,
                hora=hora,
                dia_semana=fecha_pred.weekday(),
                mes=fecha_pred.month,
                dias=1,  # predicción de 1 día
            )
            predicciones.append({
                'fecha': fecha_pred.strftime('%d/%m/%Y'),
                'kwh':   resultado['kwh'],
                'costo': resultado['costo'],
            })

        return JsonResponse({'predicciones': predicciones})

    except FileNotFoundError as e:
        return JsonResponse({'error': f'Modelo no encontrado: {e}'}, status=500)
    except Exception as e:
        return JsonResponse({'error': f'Error interno: {e}'}, status=500)

class TrabajadorListView(ListView):
    model = Trabajador
    template_name = 'Sistema_proyecto/trabajadores/lista_trabajadores.html'
    context_object_name = 'trabajadores'

class TrabajadorCreateView(CreateView):
    model = Trabajador
    form_class = TrabajadorForm
    template_name = 'Sistema_proyecto/trabajadores/form_trabajador.html'
    success_url = reverse_lazy('lista_trabajadores')

class TrabajadorUpdateView(UpdateView):
    model = Trabajador
    form_class = TrabajadorForm
    template_name = 'Sistema_proyecto/trabajadores/form_trabajador.html'
    success_url = reverse_lazy('lista_trabajadores')

class TrabajadorDeleteView(DeleteView):
    model = Trabajador
    template_name = 'Sistema_proyecto/trabajadores/confirmar_eliminar_trabajador.html'
    success_url = reverse_lazy('lista_trabajadores')

class DisponibilidadUpdateView(UpdateView):
    model = Trabajador
    form_class = DisponibilidadForm
    template_name = 'Sistema_proyecto/trabajadores/disponibilidad.html'
    success_url = reverse_lazy('lista_trabajadores')